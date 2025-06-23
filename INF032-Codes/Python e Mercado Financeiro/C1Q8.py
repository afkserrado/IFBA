'''
8. Os dados a seguir representam os preços diários de fechamentos no pregão da Bovespa entre os meses de setembro e outubro de 2019. A coluna A do Excel se refere a VALE3 (Vale do Rio Doce) e a coluna B indica GGBR4 (Gerdau). Os dados estão em um arquivo Excel na planilha denominada Plan1.

Deseja-se, então, que esses dados sejam importados para o Python com a biblioteca xlrd e que se responda aos itens a seguir.

(a) Importar os dados do Excel e transformar a coluna A em uma variável que
represente a Vale e a coluna B em outra variável que represente a coluna B.

(b) Transformar as variáveis em vetores usando a biblioteca numpy.

(c) Fazer os dois gráficos dos preços da Vale e da Gerdau usando subplot. Colocar a Vale na parte superior da figura e a Gerdau na parte inferior.

(d) Calcular os retornos das duas empresas e plotar os quatro gráficos (preço da Vale e seu retorno; preço da Gerdau e seu retorno) no formato de uma matriz com 2×2 elementos.
'''

import openpyxl as op
import numpy as np
import matplotlib.pyplot as plt

wb = op.load_workbook("Dados.xlsx") # Carrega a pasta de trabalho
plan = wb["Plan1"] # Acessa a planilha

# (a): Importando os dados da planilha
vale = [plan.cell(row = i, column = 1).value for i in range(1, plan.max_row + 1)]

gerdau = [plan.cell(row = i, column = 2).value for i in range(1, plan.max_row + 1)]

# (b): Transformando as listas em vetores
aVale = np.array(vale)
aGerdau = np.array(gerdau)

# (c): Plotando os gráficos de preço da Vale e da Gerdau
fig, matriz = plt.subplots(2, 2, figsize = (10, 8))

# Gráfico do preço da Vale
matriz[0, 0].plot(aVale, label = "VALE3", color = "orange")
matriz[0, 0].set_title("Preço do Fechamento da VALE3")
matriz[0, 0].set_xlabel("Dias")
matriz[0, 0].set_ylabel("Preço (R$)")
matriz[0, 0].legend()

# Gráfico do preço da Gerdau
matriz[0, 1].plot(aGerdau, label = "GGBR4", color = "red")
matriz[0, 1].set_title("Preço do Fechamento da GGBR4")
matriz[0, 1].set_xlabel("Dias")
matriz[0, 1].set_ylabel("Preço (R$)")
matriz[0, 1].legend()

# (d) Calcular os retornos e plotar gráficos
retornos_vale = ((aVale[1:] - aVale[:-1]) / aVale[:-1]) * 100
retornos_gerdau = ((aGerdau[1:] - aGerdau[:-1]) / aGerdau[:-1]) * 100

# Gráfico dos retornos da Vale
matriz[1, 0].plot(retornos_vale, label = "VALE3", color = "green")
matriz[1, 0].set_title("Retorno diário da VALE3")
matriz[1, 0].set_xlabel("Dias")
matriz[1, 0].set_ylabel("Retorno (%)")
matriz[1, 0].legend()

# Gráfico dos retornos da Gerdau
matriz[1, 1].plot(retornos_gerdau, label = "GGBR4", color = "blue")
matriz[1, 1].set_title("Retorno diário do GGBR4")
matriz[1, 1].set_xlabel("Dias")
matriz[1, 1].set_ylabel("Retorno (%)")
matriz[1, 1].legend()

# Plot os gráficos
plt.show()